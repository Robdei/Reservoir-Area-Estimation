import pandas as pd
import glob,os
import openpyxl
import xlsxwriter
from collections import Counter

thing=[]
for file in glob.glob("*.csv"):
	thing.append(file)
d=pd.read_csv(thing[0])
z=d['Z']
print('max:'+str(max(z)))
print('min:'+str(min(z)))
data=Counter(z)
Level=data.most_common(1)[0][0] 
thing=[]
for file in glob.glob("*.xlsx"):
	thing.append(file)
def area(level):
	x=0
	for i in z:
		if i>level:
			x=x+1
	return(x*.0005366)

if 'Splinedata.xlsx' not in thing:
	workbook = xlsxwriter.Workbook('Splinedata.xlsx')
	worksheet = workbook.add_worksheet('Spline')
	worksheet = workbook.add_worksheet('Dates')
	workbook.close()

wb=openpyxl.load_workbook('Splinedata.xlsx')
ws=wb.get_sheet_by_name('Spline')
for i in range(60):
	ws.cell(row=1,column=1).value='Reservoir Level'
	ws.cell(row=1,column=2).value='Area'
	ws.cell(row=i+2,column=1).value=Level+.25*i
	ws.cell(row=i+2,column=2).value=area(Level+.25*i)
wb.save('Splinedata.xlsx')
