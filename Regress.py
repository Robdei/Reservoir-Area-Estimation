import numpy as np
#import matplotlib.pyplot as plt
from scipy.interpolate import CubicSpline
import pandas as pd
import glob, os
from PIL import Image
import openpyxl



#-------> Change sensitivity <---------
sensitivity=71
#-------> Change sensitivity <---------




Images=[]
for file in glob.glob("*.tif"):
	Images.append(file)

d=pd.read_excel('Splinedata.xlsx')
L=d['Reservoir Level']
A=d['Area']
Level=np.fliplr([np.array(L)])[0]
Area=np.fliplr([np.array(A)])[0]
#levv=np.array(lev)
#areaa=np.array(area)
#print(area)
#print(Area)
f2 = CubicSpline(Area, Level)
#print([f2(169.4),f2(169.8),f2(171.2)])

wb=openpyxl.load_workbook('Splinedata.xlsx')
ws=wb.get_sheet_by_name('Dates')
print(['Image name','  Area  ','      Level   '])
def oceanbreeze():
	roww=0
	for image in Images:
		count=0
		roww=roww+1
		#print('working on row',roww)
		im=Image.open(image)
		w,h=im.size
		pix=im.load()
		for x in range(w):
			for y in range(h):
				if pix[x,y][1]<sensitivity and pix[x,y][1]>=65:
					count=count+1
					pix[x,y]=(130,50,20)
		ws.cell(row=roww,column=1).value=image
		ws.cell(row=roww,column=2).value=count*((15**2)/1000000)
		a=f2(count*((15**2)/1000000))
		ws.cell(row=roww,column=3).value=float(a)
		print([image,count*((15**2)/1000000),float(a)])
		im.save(image[:-4]+'check.png')

oceanbreeze()




